from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Font, Alignment
import requests
from requests import *
from configs import config
import datetime
from utils.utils import * 

def get_date_range(start_date, end_date):
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    weekdays = []
    dates = []

    current_date = start_date
    while current_date <= end_date:
        day_of_week = current_date.weekday()
        weekdays.append(day_of_week)  
        dates.append(current_date.strftime("%Y-%m-%d").split("-")[-1].lstrip('0'))
        current_date += timedelta(days=1)
    return weekdays, dates


def initHeader(start_date, end_date, file):
    weekdays, dates_header = get_date_range(start_date, end_date)
    print(f"{dates_header=}")
    dates_body = []

    weekdays_str = ['2' if day == 0 else '3' if day == 1 else '4' if day == 2 else '5' if day == 3 else '6' if day == 4 else '7' if day == 5 else 'CN' for day in weekdays]

    # handle excel
    wb = openpyxl.load_workbook(file)


    if config.SHEET_NAME in wb.sheetnames:
        ws = wb[config.SHEET_NAME]
    else:
        raise KeyError(f'Sheet {config.SHEET_NAME} does not exist.')

    day_split = start_date.split("-")
    month = f'{day_split[1].lstrip("0")}/{day_split[0]}'
    first_day = f'{day_split[-1].lstrip("0")}/{day_split[1].lstrip("0")}'
    # Tile month and first day in month
    for merged_cell in ws.merged_cells.ranges:
        if merged_cell.min_row == 4 and merged_cell.max_row == 4:  # title month
            cell = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col)
            cell.value = f"Tháng {month}"
        # first day
        if merged_cell.min_row <= 5 and merged_cell.max_row >= 6 and merged_cell.min_col == 3 and merged_cell.max_col == 3:
            cell = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col)
            cell.value = f"{first_day}"

    start_col = int(config.start_col_body)  
    start_row_date = int(config.start_row_date_header) 
    start_row_weekday = int(config.start_row_weekday_header) 
 
    fill_weekends = PatternFill(start_color="B8DBFB", end_color="B8DBFB", fill_type="solid")  

    # fill_OT = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  

    print(f"{weekdays_str=}")
    for i, weekday in enumerate(weekdays_str):
        col_letter = get_column_letter(start_col + i)

        # fill weekday
        cell_weekday = ws[f'{col_letter}{start_row_weekday}']
        cell_weekday.value = weekday
        
        # fill color weeken
        if weekday == '7' or weekday == 'CN':
            cell_weekday.fill = fill_weekends
            dates_body.append({'date': dates_header[i], 'weekend': True, 'holiday': False })
        else:
            dates_body.append({'date': dates_header[i], 'weekend': False, 'holiday': False })
             
        # fill date
        cell_date = ws[f'{col_letter}{start_row_date}']
        cell_date.value = dates_header[i]
    # save
    wb.save(file)
    print(f"Init header has been written to {file}")
    for item in dates_body:
        item['date'] = int(item['date'])
    return dates_body
        


def mergeData(dates_body, additional_info):
    merged_result = {}

    for key, values in additional_info.items():
        merged_dates_body = [dict(item) for item in dates_body]
        for value in values:
            for item in merged_dates_body:
                if item['date'] == value['date']:
                    item['type'] = value['type']
                    item['duration'] = value['duration']
        merged_result[key] = merged_dates_body
    return merged_result


def fetchData(url):
    try:
        post_response = requests.get(url)
        if post_response.status_code == 200:
            response_data = post_response.json()
            if 'data' in response_data:
                return response_data['data']
        return None
    except RequestException as e: 
        print(f"Error fetchData: {e}")
        raise SystemExit(e)



def getDataAttendence(start_date, end_date):
    headers = {'Content-type': 'application/json', 'Accept': 'application/json'}
    body_data = {
    "search": {
        "page": 0,
        "size": 2000,
        "value": "%%"
    },
    "startDate": start_date,
    "endDate": end_date,
    "type": "string"
    }
    try:
        post_response = requests.post(f"{config.API_SERVER}{config.PATH_ATTENDENCE}", json=body_data, headers=headers)
        if post_response.status_code == 200:
            response_data = post_response.json()
            if 'data' in response_data:
                employee_attendence = {}
                # print(f"payload: {response_data['data']}") 
                for attendance in response_data['data']:
                    if attendance.get('employee').get('employeeId') not in employee_attendence:
                        employee_attendence[attendance.get('employee').get('employeeId')] = []
                    employee_attendence[attendance.get('employee').get('employeeId')].append(attendance)
                return employee_attendence
        return None    
    except RequestException as e: 
        print(f"Error getDataAttendence: {e}")
        raise SystemExit(e)


def fillDataBody(dates_body, employeesHaveLeave, file):
    try:
        index = 0
        start_row_employee = 9
        wb = openpyxl.load_workbook(file)
        if config.SHEET_NAME in wb.sheetnames:
            ws = wb[config.SHEET_NAME]
        else:
            raise KeyError(f'Sheet {config.SHEET_NAME} does not exist.')

        # Fetch api get dict employee by department name
        data_response = fetchData(f"{config.API_SERVER}{config.PATH_FIND_EMPLOYEE}")

        fill_departmentNames = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
        fill_weekends = PatternFill(start_color="B8DBFB", end_color="B8DBFB", fill_type="solid")
        font_attendence = Font(name='Times New Roman', size=11, color='000000') 
        fill_bg_leave_days =  PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        font_leave_days = Font(name='Times New Roman', size=11, color='FFFFFF') 
        no_fill = PatternFill(fill_type=None)
        font_department_name = Font(name='Times New Roman', size=11, bold=True)
    
        
        # Fill body
        for department in list(data_response.keys()):
            col_letter = get_column_letter(1)

            # ws.insert_rows(start_row_employee + 1)
            # Gen and style title department name 
            cell_department_name = ws[f'{get_column_letter(1)}{start_row_employee}']
            cell_department_name.value = "P."+department
            cell_department_name.font = font_department_name
            # Fill row department name
            for col in range(1, 40):
                col_letter = get_column_letter(col)  
                cell_department = ws[f'{col_letter}{start_row_employee}']
                if col > 1: 
                    cell_department.value = None
                current_border = cell_department.border

                border_cell_department_name = Border(left=Side(border_style=None),
                    right = current_border.right if col == 40 else Side(border_style=None),
                    top = current_border.top,
                    bottom = current_border.bottom)
        
                cell_department.fill = no_fill
                cell_department.border = border_cell_department_name
               
            
            if len(data_response.get(department)) > 0:
                for employee in data_response.get(department):
                    start_row_employee = start_row_employee + 1
                    # ws.insert_rows(start_row_employee + 1)
                    index = index + 1
                    
                    cell_index = ws[f'{get_column_letter(1)}{start_row_employee}']
                    cell_index.value = index
                
                    cell_employee_name = ws[f'{get_column_letter(2)}{start_row_employee}']
                    cell_employee_name.value = employee["fullname"]
        
                    # Fill data attendence to body
                    
                    if employee['employeeId'] in employeesHaveLeave:
                        for i, employee in enumerate(employeesHaveLeave[employee['employeeId']]):
                            cell_attendence = ws[f'{get_column_letter(i + 3)}{start_row_employee}'] 
                            if 'type' in employee:
                                cell_attendence.value = employee['type']
                                cell_attendence.fill = fill_bg_leave_days
                                cell_attendence.font = font_leave_days
                            elif employee['weekend'] or employee['holiday']:
                                cell_attendence.value = ''
                                cell_attendence.fill = fill_weekends

                            else:
                                cell_attendence.value = 1
                                cell_attendence.font = font_attendence
                    else:
                        for i, date in enumerate(dates_body):
                            cell_attendence = ws[f'{get_column_letter(i + 3)}{start_row_employee}'] 
                            if date['weekend'] or date['holiday']:
                                cell_attendence.value = ''
                                cell_attendence.fill = fill_weekends
                            else:
                                cell_attendence.value = 1
                                cell_attendence.font = font_attendence


            start_row_employee = start_row_employee + 1

        wb.save(file)
        return None
    except RequestException as e: 
        print(f"Error fetchData: {e}")
        raise SystemExit(e)



def handleLeaveRequestEmployee(dict_attendence_employee, dates_body):
    result = {}
    if len(dict_attendence_employee) > 0:
        for employee in dict_attendence_employee:
            if employee not in result:
                result[employee] = []
            attendences = dict_attendence_employee[employee]
            if len(attendences) > 0:
                for attendence in attendences:
                    date = (datetime.datetime.fromtimestamp(attendence['startDate'] / 1000).date()).day
                    leave_type = attendence['leaveType']['leavetypeName']
                    duration = attendence['duration']
                    result[employee].append({'date': date, 'type': leave_type, 'duration': duration})
    
    return result




if __name__  == "__main__":

    start_date = "2024-06-01"
    end_date = "2024-06-30"

    file_download = create_excel_file(start_date, end_date)
    

    dates_body = initHeader(start_date, end_date, file_download) # dates_body list date and specical date: holiday, weekend
    # print(f'{dates_body=}')
    dict_attendence_employee = getDataAttendence("2024-06-01", "2024-06-30")
    additional_info = handleLeaveRequestEmployee(dict_attendence_employee, dates_body)
    employeesHaveLeave = mergeData(dates_body, additional_info)
    # print(f"{employeesHaveLeave=}")
    fillDataBody(dates_body, employeesHaveLeave, file_download)









